import openpyxl  # 操作excel
import config # 配置文件


class MyExcel(object):
    def __init__(self, excel_path):
        self.__excel_path = excel_path
        self.__wb = openpyxl.load_workbook(self.__excel_path, data_only=True)

    # 读数据从一个表里面拿出数据
    def get_data(self,sheet_name:str,data_positions:list):
        """从excel中的某个表格获取数据
        Args:
            sheet_name  表格名称
            data_positions  从哪几列中获得数据,第一个必须是名字
        return:
            包含每个学生信息的生成器
        """
        ws = self.__wb[sheet_name]
        for row in ws.rows:
            if row[data_positions[0]].value == None: # 如果名字为空就不管它了
                continue
            # 临时变量存放一个
            temp = []
            for index in data_positions:# 把每个同学的数据放到一个零时列表里面去
                temp.append(row[index].value)
            yield temp

    '''填家数据'''
    def add_data(self, students, sheet_name: str, name_index: str, data_positions: list):
        """将数据列表添加到excel中，列表的第一个作为索引根据
        Args:
            data_list   数据列表，第一个必须是索引根据
            sheet_name  表单的名称
            name_index      姓名所在的列
            data_positions  需要插入数据的位置
        return:
        """
        ws = self.__wb[sheet_name] # 学生表格
        for student_data in students: # 根据名字找学生
            # print(student_data)
            for name in ws[name_index]:
                if student_data[0] == name.value and name.value != '姓名': # 找到了
                    student_row = ws[name.row] # 确定这个学生所在的行
                    i = 1
                    print('开始存入{}的数据'.format(name.value))
                    for data_position in data_positions: # 把数据放入这个学生所在行的具体位置
                        try:
                            student_row[data_position].value = student_data[i]
                            i += 1
                        except Exception as ret:
                            print("产生了{}的错误，可能是因为你的存入的数据个数,不满足存放数据的位置的个数,只有前面{}个数据存入了".format(ret,len(student_data)-1))
                            continue



    def save(self):
        self.__wb.save(self.__excel_path)

def test():
    # 1. 取出信息
    MyExcel_out = MyExcel('取出.xlsx')
    generator = MyExcel_out.get_data('成绩',[1,14,16])
    # 2. 存入信息
    MyExcel_in = MyExcel('存入.xlsx')
    MyExcel_in.add_data(generator,'Sheet1','B',[29,30])
    # 3. 保存信息
    MyExcel_in.save()

def main():
    test()

if __name__ == "__main__":
    test()
